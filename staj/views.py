from django.shortcuts import render , get_object_or_404 ,HttpResponseRedirect , redirect
from django.http import *
from .models import *
from django.template import loader
from django.template.loader import get_template
from django.contrib import messages
from .forms import GunForm , OgrenciForm , StajForm , KonuForm , MulakatForm , KomisyonForm , GorusmeForm , OgrenciDosyaForm
from .utils import render_to_pdf
from openpyxl import Workbook

# notlar


# mülakat düzenleme fonksiyonu iyileştirilecek

def home(request):
    return render(request , 'home.html')



def get_ogrenci(request):

    get_ogrenci_list = Ogrenci.objects.all()
    template = loader.get_template('ogrenci/ogrenciler.html')

    query = request.GET.get('ogrenci')
    query2 = request.GET.get('tamamlandi') 

    if query:
        get_ogrenci_list = Ogrenci.objects.filter(o_no = query)
    elif query2:
        get_ogrenci_list = Ogrenci.objects.filter(staj_tamamadi_mi = True)
    context = {
        'get_ogrenci_list' : get_ogrenci_list, 
    }
    return HttpResponse(template.render(context , request))

def ogrenci_details(request , pk):

    ogrenci = get_object_or_404(Ogrenci , pk = pk)
    stajlar = Staj.objects.filter(ogrenci = ogrenci)
    context = {
        'ogrenci' : ogrenci ,
        'stajlar' : stajlar ,
    }
    return render(request , 'ogrenci/ogrenci_details.html' , context )


def add_ogrenci(request):

    if request.method == "POST":
        form = OgrenciForm(request.POST)
        if form.is_valid():
            ogrenci = form.save(commit=False)
            ogrenci.save()
            return redirect('/ogrenciler', pk=ogrenci.pk)

    else:
        form = OgrenciForm()

    return render(request , 'ogrenci/ogrenci_ekle.html' , {'form': form})

def add_ogrenci_dosya(request , pk):
    ogr = get_object_or_404(Ogrenci , pk = pk)
    if request.method == "POST":
        form = OgrenciDosyaForm(request.POST , request.FILES or None )
        if form.is_valid():
            ogrenci = form.save(commit=False)
            ogr.ogrenci_dosya = ogrenci.ogrenci_dosya
            ogr.save()
            return redirect('/ogrenciler')

    else:
        form = OgrenciDosyaForm()

    return render(request , 'ogrenci/ogrenci_dosya_ekle.html' , {'form': form})

def ogrenci_edit(request , pk):

    ogrenci = get_object_or_404(Ogrenci,pk = pk)
    
    if request.method == "POST":
        form = OgrenciForm(request.POST , instance=ogrenci)
        if form.is_valid():
            ogrenci = form.save(commit=False)
            ogrenci.save()
            return redirect('/ogrenciler' , pk=ogrenci.pk)
    else:
        form = OgrenciForm(instance=ogrenci)

    return render(request , 'ogrenci/ogrenci_edit.html' , {'form': form})
        

def ogrenci_delete(request , pk):

    ogrenci = get_object_or_404(Ogrenci,pk = pk)
    ogrenci.delete()
    return redirect('/ogrenciler')

    

def get_staj(request):
    
    template = loader.get_template('staj/stajlar.html')
    get_staj_list = Staj.objects.all()
    
    query = request.GET.get('query')
    
    if query:
        get_staj_list = Staj.objects.filter(degerlendirildi = False)

    context = {
        'get_staj_list'    : get_staj_list,
    }

    return HttpResponse(template.render(context , request))

def add_staj(request , pk):
    
    if request.method == "POST":
        form = StajForm(request.POST)

        if form.is_valid():

            staj = form.save(commit=False)
            ogr = get_object_or_404(Ogrenci , pk  = pk)
            arge = get_object_or_404(Konular , baslik = 'ARGE')

            if ogr.dgs_veya_yatay_mi and staj.onceki_staj:
                staj.ogrenci = ogr
                staj.degerlendirildi = True
                ogr.o_toplam_staj_gunu += staj.toplam_gun/2
                ogr.staj_gunu += staj.toplam_gun/2
                ogr.staj_sayisi += 1 
                staj.save()
                ogr.save()
                return redirect('/stajlar')

            if staj.sinif_durumu == '2,':
                
                if staj.toplam_gun <= 25 : 
                    staj.ogrenci = ogr
                    staj.save()
                    ogr.staj_sayisi += 1
                    ogr.staj_gunu += staj.toplam_gun
                    ogr.save()
                    return redirect('/stajlar')
                
                elif staj.toplam_gun > 25:
                    
                    if staj.konu == arge:
                        staj.ogrenci = ogr
                        staj.save()
                        ogr.staj_sayisi += 1
                        ogr.staj_gunu += staj.toplam_gun
                        ogr.save()
                        return redirect('/stajlar')
                    
                    else:
                        messages.warning(request , "2. Sınıf öğrenciler 25 Günden fazla staj günü girişi yapılamaz")
                        return redirect('/ogrenciler')


            elif ogr.staj_sayisi > 0:

                onceki_stajlar = Staj.objects.filter(ogrenci = ogr)
                
                kurum_adlari = []
                departman_adlari = []

                for i in onceki_stajlar:
                    kurum_adlari.append(i.kurum_adi)
                
                for i in onceki_stajlar:
                    departman_adlari.append(i.departman_adi)
                
                i = 0 
                
                while i < len(departman_adlari):
                    if staj.kurum_adi == kurum_adlari[i]:
                        
                        if staj.departman_adi == departman_adlari[i]:
                            messages.warning(request , "Aynı kurumun departamanında 2. kez staj yapılamaz")
                            return redirect('/ogrenciler')
                        else:  
                            staj.ogrenci = ogr
                            staj.save()
                            ogr.staj_sayisi += 1
                            ogr.staj_gunu += staj.toplam_gun
                            ogr.save()
                            return redirect('/stajlar')
                    
                    i += 1

                staj.ogrenci = ogr
                staj.save()
                ogr.staj_sayisi += 1
                ogr.staj_gunu += staj.toplam_gun
                ogr.save()
                return redirect('/stajlar')


            else:
                staj.ogrenci = ogr
                staj.save()
                ogr.staj_sayisi += 1
                ogr.staj_gunu += staj.toplam_gun
                ogr.save()
                return redirect('/stajlar')
    else:
        form = StajForm()

    return render(request , 'staj/staj_ekle.html' , {'form': form})




def staj_details(request , pk):

    staj = get_object_or_404(Staj, pk = pk)
    mulakat  = Mulakat.objects.filter(staj = staj)
    if mulakat:
        mulakat = get_object_or_404(Mulakat  , staj = staj)
        context = {
            'staj' : staj ,
            'mulakat' : mulakat,
        }
    else:
        context = {
            'staj' : staj ,
            'mulakat' : mulakat,
        }
    return render(request , 'staj/staj_details.html' , context )

def staj_edit(request,pk):
  
    staj = get_object_or_404(Staj,pk = pk)
    ogr = get_object_or_404(Ogrenci , pk  = staj.ogrenci.pk)

    if request.method == "POST":
        form = StajForm(request.POST ,instance=staj)
        if form.is_valid():

            staj = form.save(commit=False)
            arge = get_object_or_404(Konular , baslik = 'ARGE')

            if staj.sinif_durumu == '2,':
                
                if staj.toplam_gun <= 25 : 
                    staj.ogrenci = ogr
                    staj.save()
                    return redirect('/stajlar')
                
                elif staj.toplam_gun > 25:
                    
                    if staj.konu == arge:
                        staj.ogrenci = ogr
                        staj.save()
                        return redirect('/stajlar')
                    
                    else:
            
                        messages.warning(request , "25 Günden fazla staj günü girişi yapılamaz")
                        return redirect('/ogrenciler')


            elif ogr.staj_sayisi > 0:

                onceki_stajlar = Staj.objects.filter(ogrenci = ogr)
                
                kurum_adlari = []
                departman_adlari = []

                for i in onceki_stajlar:
                    kurum_adlari.append(i.kurum_adi)
                
                for i in onceki_stajlar:
                    departman_adlari.append(i.departman_adi)
                
                i = 0 
                
                while i < len(departman_adlari):
                    if staj.kurum_adi == kurum_adlari[i]:
                        
                        if staj.departman_adi == departman_adlari[i]:
                            messages.warning(request , "Bir kurumun departamanında 2. kez staj yapamazsın")
                            return redirect('/ogrenciler')
                        else:  
                            staj.ogrenci = ogr
                            staj.save()
                            return redirect('/stajlar')
                    
                    i += 1
                staj.ogrenci = ogr
                staj.save()
                return redirect('/stajlar')

            else:
                staj.ogrenci = ogr
                staj.save()
                return redirect('/stajlar')
    else:
        form = StajForm(instance=staj)

    return render(request , 'staj/staj_ekle.html' , {'form': form})


def staj_delete(request , pk):
    
    staj = get_object_or_404(Staj , pk = pk)
    ogrenci = get_object_or_404(Ogrenci , pk  = staj.ogrenci.pk)
    mulakat = Mulakat.objects.filter(staj = staj)
    gun = get_object_or_404(Gun , pk = 1)
    if mulakat:
        mulakat = get_object_or_404(Mulakat , staj = staj )

        toplamYuzde = 0.0

        toplamYuzde = (mulakat.devam * 20)*0.04 
        toplamYuzde += (mulakat.caba_calisma * 20)*0.04
        toplamYuzde += (mulakat.isi_vaktinde_yapma * 20)*0.04
        toplamYuzde += (mulakat.davranis * 20)*0.04
        toplamYuzde += (mulakat.arkadaslara_davranis * 20)*0.04
        toplamYuzde += mulakat.proje * 0.15
        toplamYuzde += mulakat.duzen * 0.05
        toplamYuzde += mulakat.sunum * 0.05
        toplamYuzde += mulakat.icerik * 0.15
        toplamYuzde += mulakat.mulakat_degerlendirmesi*0.40

        ogrenci.o_toplam_staj_gunu -= (staj.toplam_gun * toplamYuzde)/100

        if ogrenci.o_toplam_staj_gunu < gun.toplam_gun:
            ogrenci.staj_tamamadi_mi = False
    ogrenci.staj_gunu -= staj.toplam_gun
    ogrenci.staj_sayisi -= 1
    ogrenci.save()
    staj.delete()
    
    return redirect('/stajlar')

def get_konu(request):
    konular = Konular.objects.all()
    template = loader.get_template('konular/konular.html')
    stajlar = Staj.objects.all()

    query1 = request.GET.get('ilk_tarih')
    query2 = request.GET.get('son_tarih')

    if query1 and query2:
        stajlar  = Staj.objects.filter(baslama_tarihi__range=[query1,query2])
    
    konu_baslik_list = []
    konu_list = []
    iter = 0
    for i in konular:
        for j in stajlar:
            if j.konu == i:
                if i.baslik not in konu_baslik_list:
                    konu_baslik_list.append(i.baslik)
                iter +=1
        if iter != 0:
            konu_list.append(iter)              
            iter = 0
    context = {
        'konular' : konular , 
        'konu_list' : konu_list,
        'konu_baslik_list' : konu_baslik_list,
    }
    return HttpResponse(template.render(context , request))


def konu_add(request):

    if request.method == "POST":
        form = KonuForm(request.POST)
        if form.is_valid():
            konu = form.save(commit=False)
            konu.save()
            return redirect('/konular') 
    else: 
        form = KonuForm()
    
    return render (request,'konular/konu_ekle.html' , {'form' : form })

def konu_duzenle(request , pk):

    konu = get_object_or_404(Konular , pk = pk)

    if request.method == "POST":
        form = KonuForm(request.POST , instance=konu)
        if form.is_valid():
            konu = form.save(commit=False)
            konu.save()
            return redirect('/konular') 
    else: 
        form = KonuForm(instance=konu)
    
    return render (request,'konular/konu_duzenle.html' , {'form' : form })


def konu_delete(request , pk):
    konu = get_object_or_404(Konular , pk = pk)
    konu.delete()
    return redirect('/konular')


def get_mulakat(request):
    mulakatlar = Mulakat.objects.all()
    query = request.GET.get('ilk_tarih')
    query2 = request.GET.get('son_tarih')

    if query and query2 :
        mulakatlar = Mulakat.objects.filter(mulakat_tarhi__range=[query , query2])

        #bu aralıkta mulakat bulunmamaktadır uyarısı
        #mulakatlar 'if mulakatlar' şeklinde yapılabilir


    template = loader.get_template('staj/mulakatlar/mulakatlar.html')

    context = {
        'mulakatlar' : mulakatlar , 
    }

    return render(request , 'staj/mulakatlar/mulakatlar.html' , context)


def mulakat_add(request , pk):
    toplamYuzde = 0.0
    staj = get_object_or_404(Staj , pk = pk)
    ogrenci = get_object_or_404(Ogrenci , pk = staj.ogrenci.pk)
    gun = get_object_or_404(Gun , pk = 1)
    if staj.degerlendirildi == True:
        messages.warning(request , "Bu stajın mulakatı zaten yapılmışıtır , düzenlemek istiyorsanız detay sayfasına gidin")
        return redirect('/stajlar')
    try:
        gorusme = get_object_or_404(Gorusme , staj = staj)
    except:
        messages.warning(request , "Görüşme eklemeden mülakat ekleyemezsiniz")
        return redirect('/stajlar')

    if request.method == "POST":
        form = MulakatForm(request.POST)
        if form.is_valid():
            mulakat = form.save(commit=False)
            mulakat.staj = staj
            staj.degerlendirildi =  True
            gorusme.gorusme_yapıldı_mı = True
            mulakat.mulakat_tarhi = gorusme.gorusme_tarihi
            mulakat.mulakat_saati = gorusme.gorusme_saati
            gorusme.save()
            staj.save()
            mulakat.save()
            
            toplamYuzde = (mulakat.devam * 20)*0.04 
            toplamYuzde += (mulakat.caba_calisma * 20)*0.04
            toplamYuzde += (mulakat.isi_vaktinde_yapma * 20)*0.04
            toplamYuzde += (mulakat.davranis * 20)*0.04
            toplamYuzde += (mulakat.arkadaslara_davranis * 20)*0.04
            toplamYuzde += mulakat.proje * 0.15
            toplamYuzde += mulakat.duzen * 0.05
            toplamYuzde += mulakat.sunum * 0.05
            toplamYuzde += mulakat.icerik * 0.15
            toplamYuzde += mulakat.mulakat_degerlendirmesi*0.40

            ogrenci.o_toplam_staj_gunu += (staj.toplam_gun * toplamYuzde)/100
            
            if ogrenci.o_toplam_staj_gunu >= gun.toplam_gun:
                ogrenci.staj_tamamadi_mi = True
            ogrenci.save()

            return redirect('/stajlar/mulakatlar') 
    else: 
        form = MulakatForm()
    
    return render (request,'staj/mulakatlar/mulakat_ekle.html' , {'form' : form })

def mulakat_duzenle(request , pk):
    
    mulakat = get_object_or_404(Mulakat , pk = pk)

    if request.method == "POST":
        mulakat = KonuForm(request.POST , instance=mulakat)
        if form.is_valid():
            mulakat = form.save(commit=False)
            mulakat.save()
            return redirect('stajlar/mulakatlar') 
    else: 
        form = MulakatForm(instance=mulakat)
    
    return render (request,'staj/mulakatlar/mulakat_duzenle.html' , {'form' : form })


def mulakat_delete(request , pk):
    mulakat = get_object_or_404(Mulakat , pk = pk)
    staj = get_object_or_404(Staj ,  pk = mulakat.staj.pk)
    ogrenci = get_object_or_404(Ogrenci , pk = staj.ogrenci.pk)
    gun = get_object_or_404(Gun , pk = 1)
    toplamYuzde = 0.0

    toplamYuzde = (mulakat.devam * 20)*0.04 
    toplamYuzde += (mulakat.caba_calisma * 20)*0.04
    toplamYuzde += (mulakat.isi_vaktinde_yapma * 20)*0.04
    toplamYuzde += (mulakat.davranis * 20)*0.04
    toplamYuzde += (mulakat.arkadaslara_davranis * 20)*0.04
    toplamYuzde += mulakat.proje * 0.15
    toplamYuzde += mulakat.duzen * 0.05
    toplamYuzde += mulakat.sunum * 0.05
    toplamYuzde += mulakat.icerik * 0.15
    toplamYuzde += mulakat.mulakat_degerlendirmesi*0.40

    ogrenci.o_toplam_staj_gunu -= (staj.toplam_gun * toplamYuzde)/100

    if ogrenci.o_toplam_staj_gunu < gun.toplam_gun:
        ogrenci.staj_tamamadi_mi = False

    ogrenci.save()

    mulakat.delete()
    return redirect('/stajlar/mulakatlar')


def mulakat_details(request , pk):

    mulakat = get_object_or_404(Mulakat, pk = pk)
    context = {
        'mulakat' : mulakat ,
    }
    return render(request , 'staj/mulakatlar/mulakat_detail.html' , context )


def get(request):
    mulakatlar = Mulakat.objects.all()
    query = request.GET.get('ilk_tarih')
    query2 = request.GET.get('son_tarih')

    if query and query2 :
        mulakatlar = Mulakat.objects.filter(mulakat_tarhi__range=[query , query2])

        #bu aralıkta mulakat bulunmamaktadır uyarısı
        #mulakatlar 'if mulakatlar' şeklinde yapılabilir

    context = {
        'mulakatlar' : mulakatlar , 
    }

    return render_to_pdf('staj/mulakatlar/pdf2.html' , context)


def get_komisyon(request):
    komisyonlar = Komisyon.objects.all()
    template = loader.get_template('komisyon/komisyonlar.html')

    context = {
        'komisyonlar' : komisyonlar, 
    }

    return HttpResponse(template.render(context , request))


def komisyon_add(request):

    if request.method == "POST":
        form = KomisyonForm(request.POST)
        if form.is_valid():
            komisyon = form.save(commit=False)
            komisyon.save()
            return redirect('/komisyonlar') 
    else: 
        form = KomisyonForm()
    
    return render (request,'komisyon/komisyon_ekle.html' , {'form' : form })

def komisyon_duzenle(request , pk):

    komisyon = get_object_or_404(Komisyon, pk = pk)

    if request.method == "POST":
        form = KomisyonForm(request.POST , instance=komisyon)
        if form.is_valid():
            komisyon = form.save(commit=False)
            komisyon.save()
            return redirect('/komisyonlar') 
    else: 
        form = KomisyonForm(instance=komisyon)
    
    return render (request,'komisyon/komisyon_duzenle.html' , {'form' : form })


def komisyon_delete(request , pk):
    komisyon = get_object_or_404(Komisyon, pk = pk)
    komisyon.delete()
    return redirect('/komisyonlar')


def get_gorusme(request):
    gorusmeler = Gorusme.objects.filter(gorusme_yapıldı_mı = False)
    template = loader.get_template('staj/gorusme/gorusmeler.html')
    context = {
        'gorusmeler' : gorusmeler, 
    }

    return HttpResponse(template.render(context , request))


def gorusme_add(request  , pk):
    """ staj mulakat tarih bilgileri bu formdan çekilecek """
    
    staj = get_object_or_404(Staj , pk = pk)

    if request.method == "POST":
        form = GorusmeForm(request.POST)
        if form.is_valid():
            gorusme = form.save(commit=False)
            staj.gorusme_eklendi_mi = True
            gorusme.staj = staj
            gorusme.save()
            staj.save()
            return redirect('/stajlar') 
    else: 
        form = GorusmeForm()
    
    return render (request,'staj/gorusme/gorusme_ekle.html' , {'form' : form })

def gorusme_duzenle(request , pk):

    gorusme = get_object_or_404(Gorusme, pk = pk)

    if request.method == "POST":
        form = GorusmeForm(request.POST , instance=gorusme)
        if form.is_valid():
            komisyon = form.save(commit=False)
            komisyon.save()
            return redirect('/stajlar') 
    else: 
        form = GorusmeForm(instance=gorusme)
    
    return render (request,'staj/gorusme/gorusme_duzenle.html' , {'form' : form })



def get_gorusme_pdf(request):
    gorusmeler = Gorusme.objects.filter(gorusme_yapıldı_mı = False)
    context = {
        'gorusmeler' : gorusmeler , 
    }

    return render_to_pdf('staj/gorusme/pdf1.html' , context)


def gorusmeler_exel(request):
    gorusmeler = Gorusme.objects.filter(gorusme_yapıldı_mı = False)

    workbook = Workbook()
    worksheet = workbook.active

    worksheet['B1'] = 'GÖRÜŞME RAPORU'
    worksheet.merge_cells('B1:E1')

    worksheet['B3'] = 'Ad'
    worksheet['C3'] = 'Soyad'
    worksheet['D3'] = 'Görüşme Tarihi'
    worksheet['E3'] = 'Görüşme Saati'
    worksheet['F3'] = 'Komisyon Üye 1'
    worksheet['G3'] = 'Komisyon Üye 2'

    cont = 6

    for i in gorusmeler:
        worksheet.cell(row = cont , column = 2).value = i.staj.ogrenci.o_isim
        worksheet.cell(row = cont , column = 3).value = i.staj.ogrenci.o_soyisim
        worksheet.cell(row = cont , column = 4).value = i.gorusme_tarihi
        worksheet.cell(row = cont , column = 5).value = i.gorusme_saati
        worksheet.cell(row = cont , column = 6).value = i.komisyon.uye1
        worksheet.cell(row = cont , column = 7).value = i.komisyon.uye2
        
        cont += 1

    rapor = "gorusme_rapor.xlsx"
    response = HttpResponse(content_type = "application/ms-excel")
    content = "attachment ; filename = {0}".format(rapor)
    response['Content-Disposition'] = content
    workbook.save(response)
    return response

def mulakatlar_exel(request):
    mulakatlar = Mulakat.objects.all()
    query = request.GET.get('ilk_tarih')
    query2 = request.GET.get('son_tarih')

    if query and query2 :
        mulakatlar = Mulakat.objects.filter(mulakat_tarhi__range=[query , query2])

    workbook = Workbook()
    worksheet = workbook.active

    worksheet['B1'] = 'MÜLAKAT RAPORU'
    worksheet.merge_cells('B1:E1')

    worksheet['B3'] = 'Mulakat Tarihi'
    worksheet['C3'] = 'Mulakat Saati'
    worksheet['D3'] = 'Staja Devam'
    worksheet['E3'] = 'Çaba ve Çalışma'
    worksheet['F3'] = 'İşi vaktinde tamamlama'
    worksheet['G3'] = 'Amire karşı darvranış'
    worksheet['H3'] = 'İş arkadaşlarına karşı davranış'
    worksheet['I3'] = 'Proje'
    worksheet['J3'] = 'Düzen'
    worksheet['K3'] = 'Sunum'
    worksheet['L3'] = 'İçerik'
    worksheet['M3'] = 'Mülakat'

    cont = 5

    for i in mulakatlar:
        worksheet.cell(row = cont , column = 2).value = i.mulakat_tarhi
        worksheet.cell(row = cont , column = 3).value = i.mulakat_saati
        worksheet.cell(row = cont , column = 4).value = i.devam
        worksheet.cell(row = cont , column = 5).value = i.caba_calisma
        worksheet.cell(row = cont , column = 6).value = i.isi_vaktinde_yapma
        worksheet.cell(row = cont , column = 7).value = i.davranis
        worksheet.cell(row = cont , column = 8).value = i.arkadaslara_davranis
        worksheet.cell(row = cont , column = 9).value = i.proje
        worksheet.cell(row = cont , column = 10).value = i.duzen        
        worksheet.cell(row = cont , column = 11).value = i.sunum
        worksheet.cell(row = cont , column = 12).value = i.icerik
        worksheet.cell(row = cont , column = 13).value = i.mulakat_degerlendirmesi
        cont += 1

    rapor = "mulakat_rapor.xlsx"
    response = HttpResponse(content_type = "application/ms-excel")
    content = "attachment ; filename = {0}".format(rapor)
    response['Content-Disposition'] = content
    workbook.save(response)
    return response

    
def pdf_ogrenciler(request):
    ogrenciler = Ogrenci.objects.filter(staj_tamamadi_mi = True)
    context = {
        'ogrenciler' : ogrenciler , 
    }

    return render_to_pdf('ogrenci/pdf3.html' , context)

def ogrenciler_exel(request):
    ogrenciler = Ogrenci.objects.filter(staj_tamamadi_mi = True)

    workbook = Workbook()
    worksheet = workbook.active

    worksheet['B1'] = 'ÖĞRENCİ RAPORU'
    worksheet.merge_cells('B1:E1')

    worksheet['B3'] = 'Öğrenci No'
    worksheet['C3'] = 'İsim'
    worksheet['D3'] = 'Soyisim'
    worksheet['E3'] = 'Öğretim'
    worksheet['F3'] = 'Staj Durumu'

    cont = 5

    for i in ogrenciler:
        worksheet.cell(row = cont , column = 2).value = i.o_no
        worksheet.cell(row = cont , column = 3).value = i.o_isim
        worksheet.cell(row = cont , column = 4).value = i.o_soyisim
        worksheet.cell(row = cont , column = 5).value = i.o_ogretim
        worksheet.cell(row = cont , column = 6).value = i.staj_tamamadi_mi
        
        cont += 1

    rapor = "ogrenci_rapor.xlsx"
    response = HttpResponse(content_type = "application/ms-excel")
    content = "attachment ; filename = {0}".format(rapor)
    response['Content-Disposition'] = content
    workbook.save(response)
    return response


def gun_setle(request , pk):
    gun  = get_object_or_404(Gun,  pk = pk)   

    if request.method == "POST":
        form = GunForm(request.POST , instance=gun)
        if form.is_valid():
            komisyon = form.save(commit=False)
            komisyon.save()
            return redirect('/stajlar') 
    else: 
        form = GunForm(instance=gun)
    
    return render (request,'staj/gun_setle.html' , {'form' : form })